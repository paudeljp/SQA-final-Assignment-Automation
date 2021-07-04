import openpyxl
import os
# import time
from datetime import datetime
import Format_ExcelSheet

test_result_location = r'C:\Users\jeeva\PycharmProjects\QA_Assignment\Output_Result\test_result\TestResult.xlsx'

def excel_creater():
    if(os.path.exists(test_result_location)):
        workbook = openpyxl.load_workbook(test_result_location)
        worksheet = workbook['TestResults']
        return workbook,worksheet
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet('TestResults')
        worksheet.cell(row=1, column=1).value = "SN"
        worksheet.cell(row=1, column=2).value = "Test Summary"
        worksheet.cell(row=1, column=3).value = "Result"
        worksheet.cell(row=1, column=4).value = "Remarks"

        # Format_ExcelSheet.format_test_details_title(worksheet)
        workbook.save(test_result_location)
        return workbook,worksheet

def write_result(sn,test_summary,result,remarks):
    workbook,worksheet = excel_creater()
    fieldnames = (int(sn),test_summary,result,str(remarks))
    start_column = 1
    start_row = int(sn)+1
    for field in fieldnames:
        worksheet.cell(row=start_row,column=start_column).value = field
        start_column+=1
    workbook.save(test_result_location)

def write_summary(start_time, url_name):
    # workbook = openpyxl.Workbook()
    end_time = str(datetime.now())
    workbook = openpyxl.load_workbook(test_result_location)
    worksheet = workbook.create_sheet('TestSummary')
    worksheet.cell(row=1, column=1).value = "Test Executed On"
    worksheet.cell(row=1, column=2).value = start_time
    worksheet.cell(row=2, column=1).value = "Test Completed On"
    worksheet.cell(row=2, column=2).value = end_time
    worksheet.cell(row=3, column=1).value = "URL"
    worksheet.cell(row=3, column=2).value = url_name
    worksheet.cell(row=4, column=1).value = "Total Number of Test"
    worksheet.cell(row=4, column=2).value = "=(COUNTA(TestResults!A:A) - 1)"
    worksheet.cell(row=5, column=1).value = "Number of Passed Test Case"
    worksheet.cell(row=5, column=2).value = '=COUNTIF(TestResults!C:C, "PASS")'
    worksheet.cell(row=6, column=1).value = "Number of Failed Test Case"
    worksheet.cell(row=6, column=2).value = '=COUNTIF(TestResults!C:C, "FAIL")'
    worksheet.cell(row=7, column=1).value = "Number of Skipped Test Case"
    worksheet.cell(row=7, column=2).value = '=COUNTIF(TestResults!C:C, "SKIPPED")'

    Format_ExcelSheet.format_summary_title(worksheet)
    workbook.save(test_result_location)

def format_excelsheet():
    workbook = openpyxl.load_workbook(test_result_location)

    testResultworksheet = workbook['TestResults']
    Format_ExcelSheet.format_testdetails(testResultworksheet)

    testSummaryworksheet = workbook['TestSummary']
    Format_ExcelSheet.format_testsummary(testSummaryworksheet)

    workbook.save(test_result_location)

# def close_file():
    # with open(test_result_location, 'wb') as ff:
    #     os.close()
    # #
    # if (os.open(test_result_location, 777)):
    #     os.close(test_result_location, 777)
    # else:
    #     print("File is not opened already")

def remove_file():
    if (os.path.exists(test_result_location)):
        # close_file()
        os.remove(test_result_location)
    else:
        print("The file does not exist")

