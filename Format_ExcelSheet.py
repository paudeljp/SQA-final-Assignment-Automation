from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle

errorFill = PatternFill(bgColor='FF0000')
successFill = PatternFill(bgColor='00AA00')
skippedFill = PatternFill(bgColor='68A0F9')

titleFill = PatternFill(patternType='solid', fgColor='68A0F9')
darkFill = PatternFill(patternType='solid', fgColor='000000')

rightAlignment = Alignment(horizontal='right', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)

textFont = Font(bold=True, color='ffffff')

titleBorder = Border(
    left= Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

def color_summary_title(worksheet, start, end):
    for len in range(start, end + 1):
        _cell = worksheet.cell(row=len, column=1)
        _cell.fill = titleFill
        _cell.font = textFont

def format_summary_title(worksheet):
    color_summary_title(worksheet, 1, 4)
    color_summary_title(worksheet, 7, 9)
    color_summary_title(worksheet, 12, 14)


def format_test_details_title(worksheet):
    total_length = 5
    for len in range(1, total_length + 1):
        _cell = worksheet.cell(row=1, column=len)
        _cell.fill = darkFill
        _cell.font = textFont

def fit_column(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell.border = titleBorder
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 3)
        worksheet.column_dimensions[column].width = adjusted_width

def format_test_details(worksheet):

    # Success
    successRule = Rule(type="containsText", operator="containsText", formula = ['NOT(ISERROR(SEARCH("PASS",C1)))'], dxf=DifferentialStyle(fill=successFill))
    worksheet.conditional_formatting.add('C1:C200', successRule)

    # Failed
    failedRule = Rule(type="containsText", operator="containsText", formula=['NOT(ISERROR(SEARCH("FAIL",C1)))'], dxf=DifferentialStyle(fill=errorFill))
    worksheet.conditional_formatting.add('C1:C200', failedRule)

    # Skipped
    skippedRule = Rule(type="containsText", operator="containsText", formula= ['NOT(ISERROR(SEARCH("SKIPPED",C1)))'], dxf=DifferentialStyle(fill=skippedFill))
    worksheet.conditional_formatting.add('C1:C200', skippedRule)

def format_summary_details(worksheet):
    total_length = 16
    for len in range(1, total_length + 1):
        _cell = worksheet.cell(row=len, column=2)
        _cell.alignment = rightAlignment

def format_testdetails(worksheet):
    format_test_details_title(worksheet)
    format_test_details(worksheet)
    fit_column(worksheet)

def format_testsummary(worksheet):
    format_summary_title(worksheet)
    format_summary_details(worksheet)
    fit_column(worksheet)
