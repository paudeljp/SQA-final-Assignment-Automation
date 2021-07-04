import openpyxl


def writingfile(sn, test_summary,result,remarks):
    workbook, worksheet1,worksheet2 = excel_creater()
    fieldnames = (int(sn), test_summary,result,str(remarks))
    start_column = 1
    start_row = int(sn) + 1
    for fieldnames in fieldnames:
        worksheet2.cell(row=start_row, column=start_column).value = fieldname
        start_column += 1
    format_excel(worksheet2, start_row)
    fit_column(worksheet2)
    workbook.save('Output_Result/test_result/TestResult.xlsx')


def format_excel(worksheet, start_row):
    redFill = PatternFill(start_color = 'EE1111', end_color = 'EE111', fill_type='solid')
    greenFill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    blueFill = PatternFill(start_color='68A0F9', end_color='68A0F9', fill_type='solid')
    character = ('A','B','C','D')
    for ranges in character:
        cell = ranges+str(start_row)
        worksheet.conditional_formatting.add('A1:D1', FormulaRule(formula=['ISBLANK(L1)'], stopIfTrue=True,fill=blueFill))
        worksheet.conditional_formatting.add(cell, FormulaRule(formula=['=ISNUMBER(SEARCH("fail",'+cell+'))'], stopIfTrue=True, fill=redFill))
        worksheet.conditional_formatting.add(cell, FormulaRule(formula=['=ISNUMBER(SEARCH("PASS",'+ cell +'))'],stopIfTrue=True, fill=greenFill))

def fit_column(worksheet2):
        for col in worksheet2.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 0.5)
            worksheet2.columns_dimensions[column].width = adjusted_width

def excel_creater():
            fname = 'Output_Result/test_result/TestResult.xlsx'
            if(os.path.exists(fname)):
                workbook = openpyxl.load_workbook(fname)
                worksheet1 = workbook.get_sheet_by_name('Summary')
                worksheet2 = workbook.get_sheet_by_name('Details')
                return (workbook,worksheet1,worksheet2)
            else:
                workbook = openpyxl.Workbook()
                worksheet1 = workbook.create_sheet('Summary')
                worksheet2 = workbook.create_sheet('Details')
                return (workbook,worksheet1.worksheet2)

def writeheader():
            workbook, worksheet1,worksheet2 = excel_creater()
            worksheet1.cell(row=1, column=1).value = "S.No"
            worksheet1.cell(row=1, column=2).value = "Test Summary"
            worksheet1.cell(row=1, column=3).value = "Result"
            worksheet1.cell(row=1, column=4).value = "Remarks"
            workbook.save('Output_Result/test_result/TestResult')




